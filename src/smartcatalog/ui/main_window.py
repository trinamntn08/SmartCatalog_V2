# smartcatalog/ui/main_window.py
from __future__ import annotations

import threading
import io
import traceback
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
from pathlib import Path
from typing import Callable, Optional
import sqlite3
import shutil
import datetime
from PIL import Image, ImageTk

from smartcatalog.state import AppState, CatalogItem
from smartcatalog.loader.pdf_loader import build_or_update_db_from_pdf
from smartcatalog.ui.widgets.scrollable_frame import ScrollableFrame
from smartcatalog.ui.controllers.candidates_controller import CandidatesControllerMixin
from smartcatalog.ui.controllers.images_controller import ImagesControllerMixin
from smartcatalog.ui.controllers.items_controller import ItemsControllerMixin
from smartcatalog.ui.controllers.item_form_controller import ItemFormControllerMixin
from smartcatalog.loader.excel_loader import load_code_to_vi_en_from_excel, detect_excel_code_column
from smartcatalog.ui.pdf_crop_window import PdfCropWindow

import re
import hashlib
from bisect import bisect_right
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D

def _normalize_code_soft(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = s.replace("‚Äì", "-").replace("‚Äî", "-")
    s = re.sub(r"\s+", "", s)  # remove all spaces
    return s

def _normalize_header_text(s: str) -> str:
    s = str(s or "").strip().lower()
    s = s.replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _sanitize_filename(s: str) -> str:
    s = re.sub(r"[^A-Za-z0-9_-]+", "_", str(s or "").strip())
    s = s.strip("_")
    return s or "item"


def _get_image_anchor_row(img) -> Optional[int]:
    try:
        anchor = getattr(img, "anchor", None)
        if anchor is None:
            return None
        if hasattr(anchor, "_from") and getattr(anchor._from, "row", None) is not None:
            return int(anchor._from.row) + 1
        if getattr(anchor, "row", None) is not None:
            return int(anchor.row) + 1
    except Exception:
        return None
    return None


def _image_to_pil(img) -> Optional[Image.Image]:
    try:
        data = img._data()
        return Image.open(io.BytesIO(data))
    except Exception:
        pass
    try:
        ref = getattr(img, "ref", None)
        if ref:
            p = Path(ref)
            if p.exists():
                return Image.open(p)
    except Exception:
        return None
    return None

def _build_db_code_index(db_codes: list[str]) -> dict[str, str]:
    """
    normalized_code -> original_db_code
    only keep unique mappings to avoid wrong updates.
    """
    buckets: dict[str, list[str]] = {}
    for c in db_codes:
        key = _normalize_code_soft(c)
        buckets.setdefault(key, []).append(c)

    index: dict[str, str] = {}
    for k, vals in buckets.items():
        if len(vals) == 1:
            index[k] = vals[0]
    return index

def _safe_ui(root: tk.Misc, fn: Callable[[], None]) -> None:
    root.after(0, fn)


class MainWindow(
                    ttk.Frame,
                    ItemsControllerMixin,
                    CandidatesControllerMixin,
                    ImagesControllerMixin,
                    ItemFormControllerMixin,
                ):
    ...

    def __init__(self, root: tk.Tk, state: Optional[AppState] = None):
        super().__init__(root, padding=10)
        self.root = root
        self.state = state or AppState()
        self.state.ensure_dirs()

        self._sort_col: str = "id"
        self._sort_desc: bool = False

        self.status_message = tk.StringVar(value="Ch∆∞a t·∫£i d·ªØ li·ªáu")
        self._busy = tk.BooleanVar(value=False)

        # form vars
        self.var_code = tk.StringVar()
        self.var_page = tk.StringVar()

        self.var_category = tk.StringVar()
        self.var_author = tk.StringVar()
        self.var_dimension = tk.StringVar()
        self.var_small_description = tk.StringVar()
        self.var_shape = tk.StringVar()
        self.var_blade_tip = tk.StringVar()
        self.var_surface_treatment = tk.StringVar()
        self.var_material = tk.StringVar()
        self.var_validated = tk.BooleanVar(value=False)

        self._thumb_refs: list[ImageTk.PhotoImage] = []
        self._full_img_ref: Optional[ImageTk.PhotoImage] = None
        self._selected_image_path: Optional[str] = None

        self._selected: Optional[CatalogItem] = None
        

        self._build_layout()

        self._build_left_panel()
        self._build_right_panel()
        self._update_pdf_tools_label()
        self._build_status_bar()

        self.refresh_items()

    # -----------------
    # Layout
    # -----------------

    def _build_layout(self) -> None:
        self.pack(fill="both", expand=True)
        self.root.title("SmartCatalog ‚Äî Tr√¨nh qu·∫£n l√Ω danh m·ª•c")

        self.toolbar = ttk.Frame(self)
        self.toolbar.pack(fill="x", pady=(0, 8))

        self.btn_build_pdf = ttk.Button(self.toolbar, text="üìï T·∫°o/C·∫≠p nh·∫≠t CSDL t·ª´ PDF", command=self.on_choose_pdf_and_build_db)
        self.btn_build_pdf.pack(side="left", padx=(0, 6))
        
        self.btn_match_excel = ttk.Button(self.toolbar, text="C·∫≠p nh·∫≠t CSDL t·ª´ Excel", command=self.on_build_excel_db)
        self.btn_match_excel.pack(side="left")

        self.btn_backup = ttk.Button(self.toolbar, text="üíæ Backup CSDL", command=self.on_backup_data)
        self.btn_backup.pack(side="left", padx=(6, 0))

        self.btn_refresh = ttk.Button(self.toolbar, text="üîÑ L√†m m·ªõi", command=self.refresh_items)

        ttk.Separator(self.toolbar, orient="vertical").pack(side="left", fill="y", padx=8)

        self.btn_search_images  = ttk.Button(self.toolbar, text="üîç T√¨m ·∫£nh theo m√£", command=self.on_search_images_from_excel)
        self.btn_search_images .pack(side="left")

        # Panes
        self.panes = ttk.PanedWindow(self, orient="horizontal")
        self.panes.pack(side="top", fill="both", expand=True)

        self.left_pane = ttk.Frame(self.panes)
        self.right_pane = ttk.Frame(self.panes)

        self.panes.add(self.left_pane, weight=1)
        self.panes.add(self.right_pane, weight=3)

        # Scrollable container inside right pane
        self.right_scroll = ScrollableFrame(self.right_pane)
        self.right_scroll.pack(fill="both", expand=True)


    def _build_left_panel(self) -> None:
        search_frame = ttk.Frame(self.left_pane)
        search_frame.pack(fill="x", pady=(0, 6))

        ttk.Label(search_frame, text="T√¨m:").pack(side="left")
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(search_frame, textvariable=self.search_var)
        self.search_entry.pack(side="left", fill="x", expand=True, padx=6)
        self.search_entry.bind("<KeyRelease>", lambda _e: self._filter_items())

        list_frame = ttk.LabelFrame(self.left_pane, text="üì¶ S·∫£n ph·∫©m", padding=6)
        list_frame.pack(fill="both", expand=True)

        columns = (
            "id",
            "code",
            "page",
            "category",
            "author",
            "shape",
            "blade_tip",
            "dimension",
            "surface_treatment",
            "material",
            "validated",
        )
        self.items_tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=18)
        self.items_tree.heading("id", command=lambda: self._sort_by("id"))
        self.items_tree.heading("code",command=lambda: self._sort_by("code"))
        self.items_tree.heading("page", command=lambda: self._sort_by("page"))
        self.items_tree.heading("category", command=lambda: self._sort_by("category"))
        self.items_tree.heading("author", command=lambda: self._sort_by("author"))
        self.items_tree.heading("shape", command=lambda: self._sort_by("shape"))
        self.items_tree.heading("blade_tip", command=lambda: self._sort_by("blade_tip"))
        self.items_tree.heading("dimension", command=lambda: self._sort_by("dimension"))
        self.items_tree.heading("surface_treatment", command=lambda: self._sort_by("surface_treatment"))
        self.items_tree.heading("material", command=lambda: self._sort_by("material"))
        self.items_tree.heading("validated", command=lambda: self._sort_by("validated"))

        self.items_tree.column("id", width=40, anchor="center", stretch=False)
        self.items_tree.column("code", width=90, anchor="w", stretch=False)
        self.items_tree.column("page", width=40, anchor="center", stretch=False)
        self.items_tree.column("category", width=100, anchor="w", stretch=False)
        self.items_tree.column("author", width=90, anchor="w", stretch=False)
        self.items_tree.column("shape", width=80, anchor="w", stretch=False)
        self.items_tree.column("blade_tip", width=80, anchor="w", stretch=False)
        self.items_tree.column("dimension", width=80, anchor="w", stretch=False)
        self.items_tree.column("surface_treatment", width=120, anchor="w", stretch=False)
        self.items_tree.column("material", width=80, anchor="w", stretch=False)
        self.items_tree.column("validated", width=80, anchor="center", stretch=False)

        yscroll = ttk.Scrollbar(list_frame, orient="vertical", command=self.items_tree.yview)
        xscroll = ttk.Scrollbar(list_frame, orient="horizontal", command=self.items_tree.xview)
        self.items_tree.configure(yscrollcommand=yscroll.set)
        self.items_tree.configure(xscrollcommand=xscroll.set)

        # Use grid to keep scrollbars aligned with the treeview
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)
        self.items_tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")

        self.items_tree.bind("<<TreeviewSelect>>", self._on_select_item)
        self._update_sort_headers()

    def _build_right_panel(self) -> None:
        parent = self.right_scroll.inner

        self._build_item_editor_section(parent)
        self._build_images_section(parent)
        self._build_candidates_section_simple(parent)


    def _build_item_editor_section(self, parent) -> None:
        editor = ttk.LabelFrame(parent, text="üßæ S·∫£n ph·∫©m", padding=8)
        editor.pack(fill="x", pady=(0, 0))
        editor.columnconfigure(1, weight=1)

        # --- Top row: PDF info + Save button ---
        top = ttk.Frame(editor)
        top.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 8))
        top.columnconfigure(0, weight=1)

        self.pdf_tools_label = ttk.Label(top, text="Ch∆∞a ch·ªçn PDF")
        self.pdf_tools_label.grid(row=0, column=0, sticky="w")
        self.btn_save = ttk.Button(top, text="üíæ L∆∞u", command=self.on_save_item)
        self.btn_save.grid(row=0, column=1, sticky="e", padx=(0, 6))
        self.btn_add_item = ttk.Button(top, text="‚ûï Th√™m m·ªõi", command=self.on_add_item)
        self.btn_add_item.grid(row=0, column=2, sticky="e", padx=(0, 6))
        self.btn_delete_item = ttk.Button(top, text="üóëÔ∏è X√≥a", command=self.on_delete_item)
        self.btn_delete_item.grid(row=0, column=3, sticky="e")
        self.chk_validated = ttk.Checkbutton(top, text="ƒê√£ ki·ªÉm duy·ªát", variable=self.var_validated)
        self.chk_validated.grid(row=1, column=1, columnspan=3, sticky="e", pady=(4, 0))

        # --- Fields (replaces "Item fields" box) ---
        r = 1

        ttk.Label(editor, text="M√£").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_code).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="Trang").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_page).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="Ch·ªßng lo·∫°i").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_category).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="T√°c gi·∫£").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_author).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="H√¨nh d·∫°ng").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_shape).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="ƒê·∫ßu l∆∞·ª°i").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_blade_tip).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="K√≠ch th∆∞·ªõc").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_dimension).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="X·ª≠ l√Ω b·ªÅ m·∫∑t/ c√¥ng ngh·ªá").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_surface_treatment).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="Material").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_material).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="M√¥ t·∫£ t·ª´ PDF").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_small_description).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="M√¥ t·∫£ EN t·ª´ Excel").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        self.description_excel_text = scrolledtext.ScrolledText(editor, wrap="word", height=4)
        self.description_excel_text.grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="M√¥ t·∫£ VI t·ª´ Excel").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        self.description_vietnames_from_excel_text = scrolledtext.ScrolledText(editor, wrap="word", height=4)
        self.description_vietnames_from_excel_text.grid(row=r, column=1, sticky="ew", pady=3)

    def on_open_pdf_cropper(self) -> None:
        if not self.state.catalog_pdf_path:
            messagebox.showwarning("Ch∆∞a c√≥ PDF", "Vui l√≤ng t·∫°o/ch·ªçn PDF tr∆∞·ªõc.")
            return
        if not self._selected or not getattr(self._selected, "page", None):
            messagebox.showwarning("Thi·∫øu trang", "Vui l√≤ng ch·ªçn s·∫£n ph·∫©m c√≥ trang h·ª£p l·ªá.")
            return

        def after_save():
            self.refresh_items()
            self._selected = next((x for x in self.state.items_cache if x.id == self._selected.id), self._selected)
            self._reload_selected_into_form()

            if self._selected and getattr(self._selected, "page", None):
                self._render_candidates_for_page(int(self._selected.page) - 1)

            self._set_status("‚úÖ ƒê√£ l∆∞u c·∫Øt ·∫£nh v√† g·∫Øn v√†o s·∫£n ph·∫©m")


        PdfCropWindow(
            self.root,
            state=self.state,
            item_id=int(self._selected.id),
            page_1based=int(self._selected.page),
            on_after_save=after_save,
            title="C·∫Øt t·ª´ PDF",
        )


    def _update_pdf_tools_label(self) -> None:
        pdf = self.state.catalog_pdf_path
        it = self._selected

        # If label not created yet (defensive)
        if not hasattr(self, "pdf_tools_label"):
            return

        if not pdf:
            self.pdf_tools_label.configure(text="Ch∆∞a ch·ªçn PDF")
            return

        pdf_name = Path(pdf).name if not isinstance(pdf, Path) else pdf.name
        page = getattr(it, "page", None) if it else None

        if page:
            self.pdf_tools_label.configure(text=f"PDF: {pdf_name} | Trang: {page}")
        else:
            self.pdf_tools_label.configure(text=f"PDF: {pdf_name} | (ch·ªçn s·∫£n ph·∫©m)")


    def _build_images_section(self, parent) -> None:
        images_frame = ttk.LabelFrame(parent, text="üñº ·∫¢nh", padding=8)
        images_frame.pack(fill="both", expand=False, pady=(8, 0))

        thumb_container = ttk.Frame(images_frame)
        thumb_container.pack(side="left", fill="both", expand=True)

        self.thumb_canvas = tk.Canvas(thumb_container, height=180)
        self.thumb_canvas.pack(side="left", fill="both", expand=True)

        thumb_scroll = ttk.Scrollbar(thumb_container, orient="vertical", command=self.thumb_canvas.yview)
        thumb_scroll.pack(side="right", fill="y")
        self.thumb_canvas.configure(yscrollcommand=thumb_scroll.set)

        self.thumb_inner = ttk.Frame(self.thumb_canvas)
        self.thumb_canvas.create_window((0, 0), window=self.thumb_inner, anchor="nw")

        def _on_thumb_inner_configure(_e=None):
            self.thumb_canvas.configure(scrollregion=self.thumb_canvas.bbox("all"))

        self.thumb_inner.bind("<Configure>", _on_thumb_inner_configure)

        right_col = ttk.Frame(images_frame)
        right_col.pack(side="left", fill="y", padx=(10, 0))

        self.image_preview_label = ttk.Label(right_col)
        self.image_preview_label.pack(fill="both", expand=False)

        btns = ttk.Frame(right_col)
        btns.pack(fill="x", pady=(8, 0))

        ttk.Button(btns, text="‚ûï Th√™m ·∫£nh", command=self.on_add_image).pack(fill="x", pady=(0, 6))
        ttk.Button(btns, text="‚ü≥ Xoay 90¬∞", command=lambda: self.on_rotate_selected_image(90)).pack(fill="x", pady=(0, 6))
        ttk.Button(btns, text="‚ûñ X√≥a ·∫£nh ƒë√£ ch·ªçn", command=self.on_remove_selected_thumbnail).pack(fill="x")

    def _build_actions_section(self, parent) -> None:
        actions = ttk.Frame(parent)
        actions.pack(fill="x", pady=(8, 0))

        self.btn_save = ttk.Button(actions, text="üíæ L∆∞u", command=self.on_save_item)
        self.btn_save.pack(side="left", padx=(0, 6))

        self.btn_reload = None
        self.btn_clear = None
    
    #-----------------------------------------------------


    def _build_status_bar(self) -> None:
        bar = ttk.Frame(self)
        bar.pack(side="bottom", fill="x", pady=(8, 0))

        self.progress = ttk.Progressbar(bar, mode="indeterminate")
        self.progress.pack(side="left", fill="x", expand=True, padx=(0, 8))

        self.status_bar = ttk.Label(bar, textvariable=self.status_message, anchor="w")
        self.status_bar.pack(side="left")

        self._apply_busy(False)

    # -----------------
    # Busy / status
    # -----------------

    def _apply_busy(self, busy: bool) -> None:
        self._busy.set(busy)
        for w in (self.btn_build_pdf, self.btn_refresh, self.btn_match_excel, self.btn_search_images,
                  self.btn_save, self.btn_add_item, self.btn_delete_item, self.btn_backup):
            w.configure(state=("disabled" if busy else "normal"))

        if busy:
            self.progress.start(10)
        else:
            self.progress.stop()

    def _set_status(self, msg: str) -> None:
        self.status_message.set(msg)

    def _set_preview_text(self, text: str) -> None:
        """
        Legacy no-op to keep controller calls safe after removing preview widget.
        """
        return

    def _run_bg(self, title: str, work: Callable[[], None]) -> None:
        def runner():
            try:
                _safe_ui(self.root, lambda: (self._apply_busy(True), self._set_status(title)))
                work()
                _safe_ui(self.root, lambda: self._apply_busy(False))
            except Exception as exc:
                tb = traceback.format_exc()

                # ‚úÖ capture strings now
                err_text = f"{exc}\n\n{tb}"

                _safe_ui(self.root, lambda: self._apply_busy(False))
                _safe_ui(self.root, lambda: self._set_status(f"‚ùå L·ªói: {exc}"))
                _safe_ui(self.root, lambda msg=err_text: messagebox.showerror("L·ªói", msg))

        threading.Thread(target=runner, daemon=True).start()

    # -----------------
    # Actions
    # -----------------

    def on_choose_pdf_and_build_db(self) -> None:
        """
        Show existing catalog PDFs, ask whether to load a new one, then build/update DB.
        """
        # show existing catalog PDFs
        pdf_dir = self.state.data_dir / "catalog_pdfs"
        existing = []
        try:
            if pdf_dir.exists():
                existing = sorted([p.name for p in pdf_dir.glob("*.pdf")])
            if existing:
                messagebox.showinfo(
                    "PDF danh m·ª•c ƒë√£ c√≥",
                    "ƒê√£ c√≥ trong CSDL:\n" + "\n".join(existing),
                )
        except Exception:
            pass

        use_new = messagebox.askyesno(
            "T·∫£i PDF m·ªõi?",
            "B·∫°n c√≥ mu·ªën t·∫£i m·ªôt file PDF m·ªõi kh√¥ng?",
        )

        if use_new or not self.state.catalog_pdf_path:
            path = filedialog.askopenfilename(
                title="Ch·ªçn PDF danh m·ª•c",
                initialdir=str(pdf_dir) if pdf_dir.exists() else None,
                filetypes=[("T·ªáp PDF", "*.pdf"), ("T·∫•t c·∫£ t·ªáp", "*.*")],
            )
            if not path:
                return
            self.state.set_catalog_pdf(path)
            _safe_ui(self.root, self._update_pdf_tools_label)
            self._set_status(f"ƒê√£ ch·ªçn PDF: {path}")
        else:
            path = str(self.state.catalog_pdf_path)
            if not path:
                return
            run_again = messagebox.askyesno(
                "C·∫≠p nh·∫≠t l·∫°i CSDL?",
                "D√πng PDF hi·ªán t·∫°i v√† c·∫≠p nh·∫≠t l·∫°i?",
            )
            if not run_again:
                self._set_status("ƒê√£ h·ªßy c·∫≠p nh·∫≠t PDF.")
                return
            self._set_status(f"ƒêang d√πng PDF hi·ªán t·∫°i: {path}")

        # From here: we have a PDF path
        def work():
            build_or_update_db_from_pdf(self.state, None, self.status_message)
            _safe_ui(self.root, self.refresh_items)
            _safe_ui(self.root, lambda: self._set_status("‚úÖ C·∫≠p nh·∫≠t CSDL t·ª´ PDF xong"))

        self._run_bg("‚è≥ ƒêang t·∫°o/c·∫≠p nh·∫≠t CSDL t·ª´ PDF...", work)

    def on_backup_data(self) -> None:
        """
        Backup SQLite DB and assets folder to a user-chosen directory.
        """
        if not self.state or not self.state.db_path:
            messagebox.showwarning("Thi·∫øu CSDL", "Kh√¥ng t√¨m th·∫•y ƒë∆∞·ªùng d·∫´n CSDL ƒë·ªÉ backup.")
            return

        db_path = Path(self.state.db_path)
        if not db_path.exists():
            messagebox.showwarning("Thi·∫øu CSDL", f"Kh√¥ng t√¨m th·∫•y file DB: {db_path}")
            return

        dest_root = filedialog.askdirectory(
            title="Ch·ªçn th∆∞ m·ª•c l∆∞u backup",
            mustexist=True,
        )
        if not dest_root:
            return

        ts = datetime.datetime.now().strftime("%Y-%m-%d_%H%M")
        backup_dir = Path(dest_root) / f"smartcatalog_backup_{ts}"

        def work():
            try:
                backup_dir.mkdir(parents=True, exist_ok=True)

                # 1) SQLite backup (safe even if DB is open)
                backup_db_path = backup_dir / "catalog.db"
                src_conn = sqlite3.connect(str(db_path))
                try:
                    dst_conn = sqlite3.connect(str(backup_db_path))
                    try:
                        src_conn.backup(dst_conn)
                    finally:
                        dst_conn.close()
                finally:
                    src_conn.close()

                # 2) Assets backup
                assets_src = Path(self.state.assets_dir)
                assets_dst = backup_dir / "assets"
                if assets_src.exists():
                    shutil.copytree(assets_src, assets_dst, dirs_exist_ok=True)

                # 3) Catalog PDFs backup
                pdfs_src = self.state.data_dir / "catalog_pdfs"
                pdfs_dst = backup_dir / "catalog_pdfs"
                if pdfs_src.exists():
                    shutil.copytree(pdfs_src, pdfs_dst, dirs_exist_ok=True)

                # 4) Settings backup
                settings_src = Path(self.state.settings_path)
                settings_dst = backup_dir / "settings.json"
                if settings_src.exists():
                    shutil.copy2(settings_src, settings_dst)

                _safe_ui(
                    self.root,
                    lambda: messagebox.showinfo(
                        "Backup xong",
                        f"ƒê√£ backup CSDL v√† assets v√†o:\n{backup_dir}",
                    ),
                )
            except Exception as exc:
                _safe_ui(
                    self.root,
                    lambda: messagebox.showerror(
                        "Backup l·ªói",
                        f"Kh√¥ng th·ªÉ backup: {exc}",
                    ),
                )

        self._run_bg("‚è≥ ƒêang backup CSDL...", work)

    def on_build_excel_db(self) -> None:
        """
        Load an Excel file and update items.description_excel and images by matching item code.
        Matching strategy:
        1) exact code match
        2) normalized match (spaces removed, weird dashes fixed) -> only if uniquely maps to a DB code
        """
        if not self.state.db:
            messagebox.showwarning("Thi·∫øu CSDL", "Vui l√≤ng t·∫°o/t·∫£i CSDL tr∆∞·ªõc (t·ª´ PDF).")
            return

        xlsx_path = filedialog.askopenfilename(
            title="Ch·ªçn file Excel",
            filetypes=[("T·ªáp Excel", "*.xlsx *.xls"), ("T·∫•t c·∫£ t·ªáp", "*.*")],
        )
        if not xlsx_path:
            return

        def work():
            # 1) read excel (all sheets) -> {excel_code: (vi, en)}
            mapping: dict[str, tuple[str, str]] = {}
            wb = load_workbook(xlsx_path)
            sheet_names = list(wb.sheetnames)
            for sn in sheet_names:
                try:
                    m = load_code_to_vi_en_from_excel(xlsx_path, sheet_name=sn)
                except Exception:
                    continue
                for k, v in m.items():
                    key = str(k).strip()
                    if key in mapping:
                        continue
                    mapping[key] = (str(v[0]).strip(), str(v[1]).strip())

            # 2) read all DB codes once (exact + normalized index)
            conn = self.state.db.connect()
            try:
                rows = conn.execute("SELECT code FROM items").fetchall()
                db_codes = [str(r["code"]) for r in rows]
            finally:
                conn.close()

            db_code_set = set(db_codes)
            db_index = _build_db_code_index(db_codes)  # normalized -> original db code (unique only)

            # 3) build image map from Excel (embedded images)
            image_map: dict[str, list[str]] = {}
            image_rows_total = 0
            try:
                hash_to_path: dict[str, str] = {}
                per_code_hashes: dict[str, set[str]] = {}
                first_code_occurrence: dict[str, tuple[str, int]] = {}

                # Pre-pass: find first (sheet, row) occurrence of each code across all sheets
                for sn in sheet_names:
                    ws = wb[sn]
                    try:
                        _df, header_row, code_col = detect_excel_code_column(xlsx_path, sheet_name=sn)
                    except Exception:
                        continue
                    header_row_1 = header_row + 1  # openpyxl is 1-based

                    code_col_idx = None
                    for cell in ws[header_row_1]:
                        if _normalize_header_text(str(cell.value or "")) == _normalize_header_text(code_col):
                            code_col_idx = cell.column
                            break
                    if code_col_idx is None:
                        continue

                    for r in range(header_row_1 + 1, ws.max_row + 1):
                        raw_code = ws.cell(row=r, column=code_col_idx).value
                        excel_code_str = str(raw_code or "").strip()
                        if not excel_code_str:
                            continue
                        if excel_code_str not in first_code_occurrence:
                            first_code_occurrence[excel_code_str] = (sn, r)

                for sn in sheet_names:
                    ws = wb[sn]
                    try:
                        _df, header_row, code_col = detect_excel_code_column(xlsx_path, sheet_name=sn)
                    except Exception:
                        continue
                    header_row_1 = header_row + 1  # openpyxl is 1-based

                    # find code column index in header row
                    code_col_idx = None
                    for cell in ws[header_row_1]:
                        if _normalize_header_text(str(cell.value or "")) == _normalize_header_text(code_col):
                            code_col_idx = cell.column
                            break
                    if code_col_idx is None:
                        continue

                    # map row -> excel code
                    code_rows: list[int] = []
                    row_to_code: dict[int, str] = {}
                    for r in range(header_row_1 + 1, ws.max_row + 1):
                        raw_code = ws.cell(row=r, column=code_col_idx).value
                        excel_code_str = str(raw_code or "").strip()
                        if not excel_code_str:
                            continue
                        code_rows.append(r)
                        row_to_code[r] = excel_code_str
                    code_rows.sort()

                    # extract images and map to nearest code row above
                    if code_rows and getattr(ws, "_images", None):
                        out_dir = Path(self.state.assets_dir) / "excel_import"
                        out_dir.mkdir(parents=True, exist_ok=True)

                        per_code_counts: dict[str, int] = {}
                        safe_sheet = _sanitize_filename(sn)
                        for img in ws._images:
                            anchor_row = _get_image_anchor_row(img)
                            if anchor_row is None:
                                continue
                            image_rows_total += 1

                            idx = bisect_right(code_rows, anchor_row) - 1
                            if idx < 0:
                                continue
                            code_row = code_rows[idx]
                            excel_code = row_to_code.get(code_row, "")
                            if not excel_code:
                                continue
                            if first_code_occurrence.get(excel_code) != (sn, code_row):
                                continue

                            pil = _image_to_pil(img)
                            if pil is None:
                                continue

                            # dedupe by image content (hash of PNG bytes)
                            buf = io.BytesIO()
                            pil.convert("RGBA").save(buf, format="PNG")
                            data = buf.getvalue()
                            img_hash = hashlib.sha256(data).hexdigest()

                            code_hashes = per_code_hashes.setdefault(excel_code, set())
                            if img_hash in code_hashes:
                                continue
                            code_hashes.add(img_hash)

                            count = per_code_counts.get(excel_code, 0) + 1
                            per_code_counts[excel_code] = count

                            safe_code = _sanitize_filename(excel_code)
                            out_path = out_dir / f"{safe_sheet}_{safe_code}_{img_hash[:10]}.png"
                            path_str = hash_to_path.get(img_hash)
                            if path_str is None:
                                try:
                                    out_path.write_bytes(data)
                                except Exception:
                                    continue
                                path_str = str(out_path)
                                hash_to_path[img_hash] = path_str

                            lst = image_map.setdefault(excel_code, [])
                            if path_str not in lst:
                                lst.append(path_str)
            except Exception:
                image_map = {}

            total = len(mapping)
            updated = 0
            missing = 0
            missing_codes: list[str] = []
            images_updated = 0
            images_missing = 0
            i = 0

            # 4) update DB (description + images) using one connection
            conn = self.state.db.connect()
            try:
                for excel_code, desc_pair in mapping.items():
                    i += 1
                    excel_code_str = str(excel_code).strip()

                    # exact match first
                    if excel_code_str in db_code_set:
                        code_to_update = excel_code_str
                    else:
                        # normalized match (only if unique)
                        code_to_update = db_index.get(_normalize_code_soft(excel_code_str), "")

                    if code_to_update:
                        desc_vi, desc_en = desc_pair
                        cur = conn.execute(
                            "UPDATE items SET description_excel=?, description_vietnames_from_excel=? WHERE code=?",
                            (str(desc_en).strip(), str(desc_vi).strip(), code_to_update),
                        )
                        if cur.rowcount > 0:
                            updated += 1
                        else:
                            missing += 1
                            if len(missing_codes) < 30:
                                missing_codes.append(excel_code_str)
                    else:
                        missing += 1
                        if len(missing_codes) < 30:
                            missing_codes.append(excel_code_str)

                    # progress update (every 25 rows)
                    if i % 25 == 0:
                        _safe_ui(self.root, lambda i=i, total=total, updated=updated, missing=missing:
                                self._set_status(f"‚è≥ C·∫≠p nh·∫≠t Excel {i}/{total} | ƒë√£ c·∫≠p nh·∫≠t={updated} | thi·∫øu={missing}"))

                # images: link excel images into assets + item_asset_links (preferred)
                excel_asset_pdf_path = f"excel:{xlsx_path}"
                excel_asset_pdf_path_db = (
                    self.state.db.to_db_path(excel_asset_pdf_path)
                    if self.state.db
                    else excel_asset_pdf_path
                )
                for excel_code, img_paths in image_map.items():
                    # keep order but drop duplicates
                    seen: set[str] = set()
                    unique_paths: list[str] = []
                    for p in img_paths:
                        if p in seen:
                            continue
                        seen.add(p)
                        unique_paths.append(p)
                    excel_code_str = str(excel_code).strip()
                    if excel_code_str in db_code_set:
                        code_to_update = excel_code_str
                    else:
                        code_to_update = db_index.get(_normalize_code_soft(excel_code_str), "")

                    if not code_to_update:
                        images_missing += 1
                        continue

                    row = conn.execute("SELECT id FROM items WHERE code=?", (code_to_update,)).fetchone()
                    if row is None:
                        images_missing += 1
                        continue

                    item_id = int(row["id"])
                    # replace existing asset links so Excel images show in UI
                    conn.execute("DELETE FROM item_asset_links WHERE item_id=?", (item_id,))
                    conn.execute("DELETE FROM item_images WHERE item_id=?", (item_id,))
                    for idx, p in enumerate(unique_paths):
                        asset_path_db = self.state.db.to_db_path(p) if self.state.db else p
                        asset_row = conn.execute(
                            "SELECT id FROM assets WHERE pdf_path=? AND page=? AND asset_path=?",
                            (excel_asset_pdf_path_db, 0, asset_path_db),
                        ).fetchone()
                        if asset_row:
                            asset_id = int(asset_row["id"])
                        else:
                            cur = conn.execute(
                                """
                                INSERT INTO assets(pdf_path, page, asset_path, x0, y0, x1, y1, source, sha256)
                                VALUES(?,?,?,?,?,?,?,?,?)
                                """,
                                (excel_asset_pdf_path_db, 0, asset_path_db, None, None, None, None, "excel", ""),
                            )
                            asset_id = int(cur.lastrowid)
                        conn.execute(
                            """
                            INSERT OR IGNORE INTO item_asset_links(item_id, asset_id, match_method, score, verified, is_primary)
                            VALUES(?,?,?,?,?,?)
                            """,
                            (item_id, asset_id, "excel", None, 1, 1 if idx == 0 else 0),
                        )
                    images_updated += 1

                conn.commit()
            finally:
                conn.close()

            # 5) refresh UI and show summary
            _safe_ui(self.root, self.refresh_items)
            _safe_ui(self.root, lambda: self._set_status(
                f"‚úÖ Nh·∫≠p Excel xong | ƒë√£ c·∫≠p nh·∫≠t={updated} | thi·∫øu={missing} | ·∫£nh={images_updated}"
            ))
            _safe_ui(
                self.root,
                lambda: messagebox.showinfo(
                    "Nh·∫≠p Excel xong",
                    "S·ªë d√≤ng ƒë·ªçc: {total}\nƒê√£ c·∫≠p nh·∫≠t: {updated}\nM√£ thi·∫øu: {missing}\n"
                    "·∫¢nh ƒë√£ g·∫Øn: {images_updated}\n·∫¢nh thi·∫øu: {images_missing}\n"
                    "·∫¢nh t√¨m th·∫•y trong file: {image_rows_total}".format(
                        total=total,
                        updated=updated,
                        missing=missing,
                        images_updated=images_updated,
                        images_missing=images_missing,
                        image_rows_total=image_rows_total,
                    ),
                ),
            )
            if missing_codes:
                _safe_ui(
                    self.root,
                    lambda: messagebox.showwarning(
                        "M√£ thi·∫øu (m·∫´u)",
                        "M·ªôt s·ªë m√£ Excel kh√¥ng kh·ªõp v·ªõi CSDL.\n\n"
                        f"M·∫´u (t·ªëi ƒëa 30):\n" + "\n".join(missing_codes),
                    ),
                )

        self._run_bg("‚è≥ ƒêang c·∫≠p nh·∫≠t m√¥ t·∫£ v√† ·∫£nh t·ª´ Excel...", work)

    def on_search_images_from_excel(self) -> None:
        """
        Load an Excel file, match codes to DB items, and write image paths back into the same file.
        """
        if not self.state.db:
            messagebox.showwarning("Thi·∫øu CSDL", "Vui l√≤ng t·∫°o/t·∫£i CSDL tr∆∞·ªõc (t·ª´ PDF).")
            return

        xlsx_path = filedialog.askopenfilename(
            title="Ch·ªçn file Excel",
            filetypes=[("T·ªáp Excel", "*.xlsx *.xls"), ("T·∫•t c·∫£ t·ªáp", "*.*")],
        )
        if not xlsx_path:
            return
        xlsx_path = str(xlsx_path)
        export_path = str(Path(xlsx_path).with_name(f"{Path(xlsx_path).stem}_co_anh{Path(xlsx_path).suffix}"))

        def work():
            # 1) detect header + code column using existing heuristics
            _df, header_row, code_col = detect_excel_code_column(xlsx_path)

            # 2) build DB code indexes (same logic as on_build_excel_db)
            conn = self.state.db.connect()
            try:
                rows = conn.execute("SELECT code FROM items").fetchall()
                db_codes = [str(r["code"]) for r in rows]
            finally:
                conn.close()

            db_code_set = set(db_codes)
            db_index = _build_db_code_index(db_codes)  # normalized -> original db code (unique only)

            items = self.state.db.list_items()
            code_to_images: dict[str, list[str]] = {str(it.code): list(it.images or []) for it in items}

            # 3) update Excel in-place (preserve layout)
            wb = load_workbook(xlsx_path)
            ws = wb.active
            header_row_1 = header_row + 1  # openpyxl is 1-based

            # find code column index in header row
            code_col_idx = None
            for cell in ws[header_row_1]:
                if _normalize_header_text(str(cell.value or "")) == _normalize_header_text(code_col):
                    code_col_idx = cell.column
                    break
            if code_col_idx is None:
                raise ValueError(f"Kh√¥ng t√¨m th·∫•y c·ªôt m√£ '{code_col}' trong d√≤ng ti√™u ƒë·ªÅ Excel.")

            # write rows
            updated = 0
            total = 0
            matched = 0
            sample_excel_codes: list[str] = []
            rows_with_images: list[tuple[int, list[str]]] = []
            for r in range(header_row_1 + 1, ws.max_row + 1):
                raw_code = ws.cell(row=r, column=code_col_idx).value
                excel_code_str = str(raw_code or "").strip()
                if not excel_code_str:
                    continue
                if len(sample_excel_codes) < 5:
                    sample_excel_codes.append(excel_code_str)
                total += 1

                if excel_code_str in db_code_set:
                    code_to_match = excel_code_str
                else:
                    code_to_match = db_index.get(_normalize_code_soft(excel_code_str), "")

                imgs = code_to_images.get(code_to_match, []) if code_to_match else []
                if code_to_match:
                    matched += 1
                if imgs:
                    updated += 1
                    rows_with_images.append((r, imgs))

            # Insert image rows (bottom-up to keep indexes stable)
            if rows_with_images:
                px_to_emu = 9525
                pad = 6

                for r, imgs in rows_with_images:
                    img_row = r + 1  # row below code
                    ws.merge_cells(start_row=img_row, start_column=1, end_row=img_row, end_column=4)

                    def col_width_px(col_idx: int) -> int:
                        letter = get_column_letter(col_idx)
                        w = ws.column_dimensions[letter].width
                        if w is None:
                            w = ws.column_dimensions["A"].width or 8.43
                        # Excel column width to pixels (approx)
                        return int(w * 7 + 5)

                    # Load images and compute base sizes
                    loaded: list[tuple[Image.Image, int, int]] = []
                    for img_path in imgs:
                        try:
                            p = Path(img_path)
                            if not p.exists():
                                continue
                            pil = Image.open(p).convert("RGBA")
                            # Rotate to landscape if needed
                            if pil.height > pil.width:
                                pil = pil.rotate(90, expand=True)
                            w, h = pil.size
                            if h <= 0 or w <= 0:
                                continue
                            loaded.append((pil, w, h))
                        except Exception:
                            continue

                    if not loaded:
                        continue

                    # Order images: small (left) -> big (right)
                    loaded.sort(key=lambda t: t[1] * t[2])

                    # Use original sizes; keep existing column widths and center images
                    total_w = sum(w for _, w, _ in loaded) + pad * max(0, len(loaded) - 1)
                    max_h = max(h for _, _, h in loaded)

                    # Set row height (points). Approx: 1 pt ~= 1.333 px
                    min_h = max_h + 20
                    ws.row_dimensions[img_row].height = max(min_h, max_h + 6) / 1.333

                    available_w = sum(col_width_px(c) for c in range(1, 5))
                    if available_w < 1:
                        available_w = total_w
                    x_off = max(0, int((available_w - total_w) / 2))
                    col_widths = [col_width_px(c) for c in range(1, 5)]
                    v_pad = 10
                    row_height_px = int((ws.row_dimensions[img_row].height or 0) * 1.333)
                    target_h = max(1, row_height_px - (v_pad * 2))

                    # Global scale to fit both row height and merged column width
                    scale_h = 1.0 if max_h <= target_h else (target_h / float(max_h))
                    scale_w = 1.0 if total_w <= available_w else (available_w / float(total_w))
                    scale = min(1.0, scale_h, scale_w)

                    for pil, w, h in loaded:
                        try:
                            new_w = max(1, int(w * scale))
                            new_h = max(1, int(h * scale))

                            buf = io.BytesIO()
                            if new_w != w or new_h != h:
                                pil_resized = pil.resize((new_w, new_h), Image.LANCZOS)
                            else:
                                pil_resized = pil
                            pil_resized.save(buf, format="PNG")
                            buf.seek(0)
                            xl_img = XLImage(buf)
                            xl_img.width = new_w
                            xl_img.height = new_h

                            # translate x_off into (column, colOff)
                            col_idx = 0
                            col_off = x_off
                            while col_idx < len(col_widths) - 1 and col_off >= col_widths[col_idx]:
                                col_off -= col_widths[col_idx]
                                col_idx += 1

                            row_height_px = int((ws.row_dimensions[img_row].height or 0) * 1.333)
                            y_off = 0
                            if row_height_px > new_h + v_pad * 2:
                                y_off = int((row_height_px - new_h) / 2)
                            elif row_height_px > new_h:
                                y_off = v_pad

                            marker = AnchorMarker(
                                col=col_idx,
                                colOff=int(col_off * px_to_emu),
                                row=img_row - 1,
                                rowOff=int(y_off * px_to_emu),
                            )
                            ext = XDRPositiveSize2D(new_w * px_to_emu, new_h * px_to_emu)
                            xl_img.anchor = OneCellAnchor(_from=marker, ext=ext)
                            ws.add_image(xl_img)

                            x_off += new_w + pad
                        except Exception:
                            continue

            wb.save(export_path)

            if matched == 0:
                sample_db_codes = db_codes[:5]
                _safe_ui(
                    self.root,
                    lambda: messagebox.showwarning(
                        "Kh√¥ng c√≥ kh·ªõp",
                        "Kh√¥ng c√≥ m√£ Excel n√†o kh·ªõp v·ªõi m√£ trong CSDL.\n\n"
                        f"C·ªôt m√£ ph√°t hi·ªán: {code_col}\n"
                        f"D√≤ng ti√™u ƒë·ªÅ: {header_row_1}\n"
                        f"M·∫´u m√£ Excel: {sample_excel_codes}\n"
                        f"M·∫´u m√£ CSDL: {sample_db_codes}",
                    ),
                )

            _safe_ui(self.root, lambda: messagebox.showinfo(
                "Xu·∫•t file xong",
                f"M√£ kh·ªõp: {matched}/{total}\nD√≤ng c√≥ ·∫£nh: {updated}/{total}\nƒê√£ l∆∞u: {export_path}",
            ))
            _safe_ui(
                self.root,
                lambda: self._set_status(f"‚úÖ Xu·∫•t ·∫£nh ra Excel: kh·ªõp {matched}/{total}, ·∫£nh {updated}/{total}")
            )

        self._run_bg("‚è≥ Extracting images by code...", work)


###################################################################################################
def create_main_window(root: tk.Tk, state: Optional[AppState] = None) -> MainWindow:
    return MainWindow(root, state=state)
